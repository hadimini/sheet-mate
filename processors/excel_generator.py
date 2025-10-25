import calendar
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import tempfile

logger = logging.getLogger(__name__)


class TimeSheetGenerator:
    def __init__(self) -> None:
        self.current_date = datetime.now()
        self.month = self.current_date.month
        self.year = self.current_date.year

    async def generate_timesheet(self, employee_name: str) -> str:
        """Generates a clean, simple timesheet Excel file"""
        try:
            logger.info(f'📊 Generating timesheet for {employee_name}...')

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = 'Timesheet'

            # SIMPLE HEADER
            ws['A1'] = f'Timesheet - {employee_name}'
            ws['A1'].font = Font(size=14, bold=True)

            ws['A2'] = f'Period: {calendar.month_name[self.month]} {self.year}'
            ws['A2'].font = Font(bold=True)

            headers = [
                'Regular hours', 'Overtime hours', 'Vacation hours', 'Sick hours', 'Holiday hours'
            ]

            for row, header in enumerate(headers, 6):
                cell = ws.cell(row=row, column=1, value=header)
                cell.font = Font(bold=True)

            # FILL DATES FOR THE MONTH
            total_days = calendar.monthrange(self.year, self.month)[1]
            day_names = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']

            for day in range(1, total_days + 1):
                align_center = Alignment(horizontal='center')
                font_bold = Font(bold=True)
                col_offset = 4
                current_date = datetime(self.year, self.month, day)
                day_name = day_names[current_date.weekday()]
                is_weekend = current_date.weekday() >= 5

                # Date
                day_number_cell = ws.cell(row=4, column=col_offset + day, value=current_date.strftime('%d'))
                day_number_cell.alignment = align_center
                day_number_cell.font = font_bold
                day_name_cell = ws.cell(row=5, column=col_offset + day, value=day_name)
                day_name_cell.alignment = align_center
                day_name_cell.font = font_bold

                if is_weekend:
                    fill = PatternFill(fill_type='solid', fgColor='808080')
                    day_number_cell.fill = fill
                    day_name_cell.fill = fill

                    for r in range(6, 11):
                        ws.cell(row=r, column=col_offset + day).fill = fill

            with tempfile.NamedTemporaryFile(
                    prefix=f'Time_sheet_{employee_name.replace(" ", "_")}_{self.month}_{self.year}_',
                    suffix='.xlsx'
            ) as tmp_file:
                file_path = tmp_file.name

            wb.save(file_path)
            logger.info(f'✅ Timesheet saved: {file_path}')

            return file_path

        except Exception as e:
            logger.error(f'❌ Error: {str(e)}')
            import traceback
            logger.error(f'🔍 Details: {traceback.format_exc()}')
            raise